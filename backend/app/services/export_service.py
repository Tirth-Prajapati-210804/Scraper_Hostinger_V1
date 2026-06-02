from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO
import re
from statistics import mean

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.core.logging import get_logger
from app.models.all_flight_result import AllFlightResult
from app.models.route_group import RouteGroup

log = get_logger(__name__)
_MISSING_VALUE = "N-A"
_INVALID_SHEET_TITLE_RE = re.compile(r"[\[\]:*?/\\]")

_MAIN_HEADERS = [
    "Date",
    "Dep Airport",
    "Arrival Airport",
    "Nights",
    "Airline",
    "Stop Result",
    "Duration",
    "Flight Price",
]

_MULTI_CITY_HEADERS = [
    "Date",
    "Ending Date",
    "Dep Airport",
    "Arrival Airport",
    "Return From",
    "Nights",
    "Airline",
    "Stop Result",
    "Duration",
    "Flight Price",
]

_DEALS_HEADERS = [
    "Rank",
    "Origin",
    "Destination",
    "Date",
    "Airline",
    "Price",
    "Savings vs Avg",
]

_SUMMARY_HEADERS = [
    "Origin",
    "Records",
    "Lowest Price",
    "Average Price",
]

_WEEKEND_HEADERS = [
    "Origin",
    "Destination",
    "Date",
    "Airline",
    "Price",
]


def _safe_stop_label(value: object, stops: object = None) -> str:
    if isinstance(value, str) and value.strip():
        return value.strip()
    if isinstance(stops, (int, float)):
        stop_count = int(stops)
        if stop_count <= 0:
            return "Direct"
        if stop_count == 1:
            return "1 Stop"
        return f"{stop_count} Stops"
    return ""


def _format_duration_minutes(minutes: int) -> str:
    hours, mins = divmod(minutes, 60)
    return f"{hours}h {mins}m"


def _parse_duration_text(value: object) -> int | None:
    if not isinstance(value, str) or not value.strip():
        return None
    import re

    text = value.lower().replace("hours", "h").replace("hour", "h")
    text = text.replace("minutes", "m").replace("minute", "m").replace("mins", "m").replace("min", "m")
    hours_match = re.search(r"(\d+)\s*h", text)
    mins_match = re.search(r"(\d+)\s*m", text)
    hours = int(hours_match.group(1)) if hours_match else 0
    mins = int(mins_match.group(1)) if mins_match else 0
    total = hours * 60 + mins
    return total if total > 0 else None


def _duration_label_from_minutes(values: list[int]) -> str:
    return " / ".join(_format_duration_minutes(value) for value in values if value > 0)


def _safe_duration_label(result: AllFlightResult) -> str:
    itinerary = getattr(result, "itinerary_data", None)
    if isinstance(itinerary, dict):
        raw_durations = itinerary.get("leg_durations")
        if isinstance(raw_durations, list):
            durations = [int(value) for value in raw_durations if isinstance(value, (int, float)) and int(value) > 0]
            if durations:
                return _duration_label_from_minutes(durations)

        legs = itinerary.get("legs")
        if isinstance(legs, list):
            durations: list[int] = []
            for leg in legs:
                if not isinstance(leg, dict):
                    continue
                raw_minutes = leg.get("duration_minutes")
                if isinstance(raw_minutes, (int, float)) and int(raw_minutes) > 0:
                    durations.append(int(raw_minutes))
                    continue
                parsed = _parse_duration_text(leg.get("duration_text"))
                if parsed:
                    durations.append(parsed)
            if durations:
                return _duration_label_from_minutes(durations)

        parsed_parts = [
            _parse_duration_text(part)
            for part in str(itinerary.get("duration_text") or "").split("/")
        ]
        durations = [value for value in parsed_parts if value]
        if durations:
            return _duration_label_from_minutes(durations)

    duration = getattr(result, "duration_minutes", None)
    if isinstance(duration, int) and duration > 0:
        return _format_duration_minutes(duration)
    return ""


def _duration_rank(result: AllFlightResult) -> int:
    duration = getattr(result, "duration_minutes", None)
    return duration if isinstance(duration, int) and duration > 0 else 10**9


def _stops_rank(result: AllFlightResult) -> int:
    stops = getattr(result, "stops", None)
    return stops if isinstance(stops, int) and stops >= 0 else 10**9


def _result_sort_key(result: AllFlightResult) -> tuple[float, int, int]:
    return (float(result.price), _duration_rank(result), _stops_rank(result))


def _set_date_cell(ws, *, row: int, column: int, value: object):
    if isinstance(value, str):
        try:
            value = date.fromisoformat(value)
        except ValueError:
            pass
    cell = ws.cell(row=row, column=column, value=value)
    cell.number_format = "DD-MM-YYYY"
    return cell


def _safe_sheet_title(wb: Workbook, value: object, *, fallback: str = "Sheet") -> str:
    base = _INVALID_SHEET_TITLE_RE.sub("-", str(value or "").strip()).strip("' ")
    if not base:
        base = fallback
    base = base[:31]
    if base not in wb.sheetnames:
        return base

    suffix = 2
    while True:
        suffix_text = f"-{suffix}"
        candidate = f"{base[:31 - len(suffix_text)]}{suffix_text}"
        if candidate not in wb.sheetnames:
            return candidate
        suffix += 1


def _export_dates(route_group: RouteGroup, fallback_dates: list[date]) -> list[date]:
    unique_fallback = sorted({d for d in fallback_dates if isinstance(d, date)})
    configured_start = getattr(route_group, "start_date", None)
    configured_end = getattr(route_group, "end_date", None)
    raw_days_ahead = getattr(route_group, "days_ahead", None)

    configured_start = configured_start if isinstance(configured_start, date) else None
    configured_end = configured_end if isinstance(configured_end, date) else None
    days_ahead = max(1, min(raw_days_ahead, 730)) if isinstance(raw_days_ahead, int) else None

    if configured_start or configured_end:
        if configured_start is None:
            configured_start = unique_fallback[0] if unique_fallback else date.today()
        if configured_end is None:
            configured_end = configured_start + timedelta(days=(days_ahead or 1) - 1)
        if configured_end >= configured_start:
            total_days = min((configured_end - configured_start).days + 1, 730)
            return [configured_start + timedelta(days=i) for i in range(total_days)]

    return unique_fallback


def export_route_group(
    route_group: RouteGroup,
    all_results: list[AllFlightResult],
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    if not all_results:
        ws = wb.create_sheet("No Data")
        ws["A1"] = "No results available"
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    if route_group.trip_type == "multi_city":
        return _export_multi_city_route_group(wb, route_group, all_results)

    # --------------------------------------------------
    # LOOKUPS
    # --------------------------------------------------

    all_dates = _export_dates(route_group, [r.depart_date for r in all_results])

    cheapest_by_origin_date: dict[tuple[str, object], AllFlightResult] = {}
    prices_by_route: dict[tuple[str, str], list[float]] = {}

    for r in all_results:
        key = (r.origin, r.depart_date)

        if key not in cheapest_by_origin_date:
            cheapest_by_origin_date[key] = r
        elif _result_sort_key(r) < _result_sort_key(cheapest_by_origin_date[key]):
            cheapest_by_origin_date[key] = r

        route_key = (r.origin, r.destination)
        prices_by_route.setdefault(route_key, []).append(float(r.price))

    # --------------------------------------------------
    # MAIN ORIGIN SHEETS
    # --------------------------------------------------

    sheet_name_map = route_group.sheet_name_map or {
        o: o for o in route_group.origins
    }

    for origin, sheet_name in sheet_name_map.items():
        ws = wb.create_sheet(title=_safe_sheet_title(wb, sheet_name, fallback=origin))
        _write_header_row(ws, _MAIN_HEADERS)

        for row_idx, d in enumerate(all_dates, start=2):
            result = cheapest_by_origin_date.get((origin, d))

            _set_date_cell(ws, row=row_idx, column=1, value=d)
            ws.cell(row=row_idx, column=2, value=origin)
            ws.cell(
                row=row_idx,
                column=3,
                value=route_group.destination_label,
            )
            ws.cell(row=row_idx, column=4, value=route_group.nights)

            if result:
                ws.cell(row=row_idx, column=5, value=result.airline)
                ws.cell(
                    row=row_idx,
                    column=6,
                    value=_safe_stop_label(result.stop_label, result.stops),
                )
                ws.cell(
                    row=row_idx,
                    column=7,
                    value=_safe_duration_label(result),
                )
                ws.cell(
                    row=row_idx,
                    column=8,
                    value=int(round(float(result.price))),
                )
            else:
                ws.cell(row=row_idx, column=5, value=_MISSING_VALUE)
                ws.cell(row=row_idx, column=6, value=_MISSING_VALUE)
                ws.cell(row=row_idx, column=7, value=_MISSING_VALUE)
                ws.cell(row=row_idx, column=8, value=_MISSING_VALUE)

        _autosize_columns(ws)

    # --------------------------------------------------
    # SPECIAL JOURNEY SHEETS  (additional return / multi-city legs the
    # operator added in the "Advanced Routes" form)
    # --------------------------------------------------

    for sheet in route_group.special_sheets or []:
        sheet_name = sheet.get("name") or "Journey"
        sheet_origin = (sheet.get("origin") or "").upper()
        sheet_dest_label = sheet.get("destination_label") or sheet_origin
        sheet_dests = [d.upper() for d in (sheet.get("destinations") or [])]
        columns = int(sheet.get("columns") or 4)

        ws = wb.create_sheet(title=_safe_sheet_title(wb, sheet_name, fallback="Journey"))

        if columns >= 6:
            _write_header_row(ws, _MAIN_HEADERS)
        else:
            _write_header_row(
                ws, ["Date", "Dep Airport", "Arrival Airport", "Flight Price"]
            )

        # cheapest result per date across this special sheet's destinations
        cheapest_per_date: dict[object, AllFlightResult] = {}
        for r in all_results:
            if r.origin != sheet_origin or r.destination not in sheet_dests:
                continue
            if (
                r.depart_date not in cheapest_per_date
                or _result_sort_key(r) < _result_sort_key(cheapest_per_date[r.depart_date])
            ):
                cheapest_per_date[r.depart_date] = r

        for row_idx, d in enumerate(all_dates, start=2):
            result = cheapest_per_date.get(d)

            _set_date_cell(ws, row=row_idx, column=1, value=d)
            ws.cell(row=row_idx, column=2, value=sheet_origin)
            ws.cell(row=row_idx, column=3, value=sheet_dest_label)

            if columns >= 6:
                ws.cell(row=row_idx, column=4, value=route_group.nights)
                if result:
                    ws.cell(row=row_idx, column=5, value=result.airline)
                    ws.cell(row=row_idx, column=6, value=_safe_stop_label(result.stop_label, result.stops))
                    ws.cell(row=row_idx, column=7, value=_safe_duration_label(result))
                    ws.cell(
                        row=row_idx,
                        column=8,
                        value=int(round(float(result.price))),
                    )
                else:
                    ws.cell(row=row_idx, column=5, value=_MISSING_VALUE)
                    ws.cell(row=row_idx, column=6, value=_MISSING_VALUE)
                    ws.cell(row=row_idx, column=7, value=_MISSING_VALUE)
                    ws.cell(row=row_idx, column=8, value=_MISSING_VALUE)
            else:
                if result:
                    ws.cell(
                        row=row_idx,
                        column=4,
                        value=int(round(float(result.price))),
                    )
                else:
                    ws.cell(row=row_idx, column=4, value=_MISSING_VALUE)

        _autosize_columns(ws)

    # --------------------------------------------------
    # BEST DEALS SHEET
    # --------------------------------------------------

    deals = []

    for r in all_results:
        route_key = (r.origin, r.destination)
        avg_price = mean(prices_by_route[route_key])

        savings = avg_price - float(r.price)

        deals.append(
            {
                "origin": r.origin,
                "destination": r.destination,
                "date": r.depart_date,
                "airline": r.airline,
                "price": float(r.price),
                "savings": savings,
            }
        )

    deals.sort(
        key=lambda x: (
            -x["savings"],
            x["price"],
        )
    )

    ws = wb.create_sheet("Best Deals")
    _write_header_row(ws, _DEALS_HEADERS)

    for i, d in enumerate(deals[:25], start=2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=d["origin"])
        ws.cell(row=i, column=3, value=d["destination"])
        _set_date_cell(ws, row=i, column=4, value=d["date"])
        ws.cell(row=i, column=5, value=d["airline"])
        ws.cell(row=i, column=6, value=int(round(d["price"])))
        ws.cell(row=i, column=7, value=int(round(d["savings"])))

    _autosize_columns(ws)

    # --------------------------------------------------
    # WEEKEND DEALS
    # --------------------------------------------------

    weekend = [
        r for r in all_results
        if r.depart_date.weekday() in (4, 5, 6)
    ]

    weekend.sort(key=_result_sort_key)

    ws = wb.create_sheet("Weekend Deals")
    _write_header_row(ws, _WEEKEND_HEADERS)

    for i, r in enumerate(weekend[:25], start=2):
        ws.cell(row=i, column=1, value=r.origin)
        ws.cell(row=i, column=2, value=r.destination)
        _set_date_cell(ws, row=i, column=3, value=r.depart_date)
        ws.cell(row=i, column=4, value=r.airline)
        ws.cell(row=i, column=5, value=int(round(float(r.price))))

    _autosize_columns(ws)

    # --------------------------------------------------
    # ORIGIN SUMMARY
    # --------------------------------------------------

    ws = wb.create_sheet("Summary")
    _write_header_row(ws, _SUMMARY_HEADERS)

    row = 2

    for origin in route_group.origins:
        rows = [r for r in all_results if r.origin == origin]

        if not rows:
            continue

        prices = [float(r.price) for r in rows]

        ws.cell(row=row, column=1, value=origin)
        ws.cell(row=row, column=2, value=len(rows))
        ws.cell(row=row, column=3, value=int(round(min(prices))))
        ws.cell(row=row, column=4, value=int(round(mean(prices))))

        row += 1

    _autosize_columns(ws)

    # --------------------------------------------------
    # FINISH
    # --------------------------------------------------

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.read()


def _export_multi_city_route_group(
    wb: Workbook,
    route_group: RouteGroup,
    all_results: list[AllFlightResult],
) -> bytes:
    itinerary_rows = [r for r in all_results if r.itinerary_data]
    if not itinerary_rows:
        ws = wb.create_sheet("No Data")
        ws["A1"] = "No itinerary results available"
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    sheet_name_map = route_group.sheet_name_map or {o: o for o in route_group.origins}
    all_dates = _export_dates(route_group, [row.depart_date for row in itinerary_rows])

    cheapest_by_route_date: dict[tuple[str, str, object], AllFlightResult] = {}
    for row in itinerary_rows:
        key = (row.origin, row.destination, row.depart_date)
        current = cheapest_by_route_date.get(key)
        if current is None or _result_sort_key(row) < _result_sort_key(current):
            cheapest_by_route_date[key] = row

    rows_by_route: dict[tuple[str, str], list[AllFlightResult]] = {}
    for row in cheapest_by_route_date.values():
        rows_by_route.setdefault((row.origin, row.destination), []).append(row)

    itinerary_prices_by_origin: dict[str, list[float]] = {}
    all_itinerary_prices: list[AllFlightResult] = []

    origin_destination_counts: dict[str, int] = {}
    for origin, _destination in rows_by_route:
        origin_destination_counts[origin] = origin_destination_counts.get(origin, 0) + 1

    def build_sheet_name(origin: str, destination: str) -> str:
        base_name = sheet_name_map.get(origin, origin)
        if origin_destination_counts.get(origin, 0) > 1:
            base_name = f"{base_name}-{destination}"

        return _safe_sheet_title(wb, base_name, fallback=f"{origin}-{destination}")

    for (origin, destination), rows in sorted(rows_by_route.items()):
        ws = wb.create_sheet(title=build_sheet_name(origin, destination))
        _write_header_row(ws, _MULTI_CITY_HEADERS)

        rows_by_date = {row.depart_date: row for row in rows}

        for row_idx, depart_date in enumerate(all_dates, start=2):
            result = rows_by_date.get(depart_date)
            itinerary = result.itinerary_data if isinstance(result, object) and result else {}
            if not isinstance(itinerary, dict):
                itinerary = {}
            return_date = itinerary.get("return_date")
            # Return-leg origin: prefer the actual airport flown (e.g. FCO when the
            # group searched the ROM metro code), else the searched return origin.
            return_from = (
                itinerary.get("actual_return_origin")
                or itinerary.get("return_origin")
                or (itinerary.get("inbound") or {}).get("origin")
            )

            _set_date_cell(ws, row=row_idx, column=1, value=depart_date)
            if return_date:
                _set_date_cell(ws, row=row_idx, column=2, value=return_date)
            else:
                ws.cell(row=row_idx, column=2, value=_MISSING_VALUE)
            ws.cell(row=row_idx, column=3, value=origin)
            ws.cell(row=row_idx, column=4, value=destination)
            ws.cell(row=row_idx, column=5, value=return_from or _MISSING_VALUE)
            ws.cell(row=row_idx, column=6, value=route_group.nights)
            if result:
                ws.cell(row=row_idx, column=7, value=result.airline)
                ws.cell(row=row_idx, column=8, value=_safe_stop_label(result.stop_label, result.stops))
                ws.cell(row=row_idx, column=9, value=_safe_duration_label(result))
                ws.cell(row=row_idx, column=10, value=int(round(float(result.price))))
                itinerary_prices_by_origin.setdefault(origin, []).append(float(result.price))
                all_itinerary_prices.append(result)
            else:
                ws.cell(row=row_idx, column=7, value=_MISSING_VALUE)
                ws.cell(row=row_idx, column=8, value=_MISSING_VALUE)
                ws.cell(row=row_idx, column=9, value=_MISSING_VALUE)
                ws.cell(row=row_idx, column=10, value=_MISSING_VALUE)

        _autosize_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _write_header_row(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def _autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        max_length = max(
            (
                len(str(c.value))
                for c in col_cells
                if c.value is not None
            ),
            default=0,
        )

        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max_length + 3
