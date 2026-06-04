from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.providers.base import ProviderResult
from app.services.export_service import export_route_group
from app.services.price_collector import PriceCollector


def test_multi_city_export_uses_itinerary_sheet_shape() -> None:
    group = SimpleNamespace(
        trip_type="multi_city",
        origins=["YYZ"],
        nights=11,
        sheet_name_map={"YYZ": "Toronto Open Jaw"},
    )

    results = [
        SimpleNamespace(
            origin="YYZ",
            destination="BER",
            depart_date=date(2026, 5, 20),
            airline="Icelandair / Lufthansa",
            price=829.0,
            stops=1,
            stop_label="1 Stop",
            itinerary_data={
                "return_date": "2026-05-31",
                "return_origin": "BUD",
                "outbound_airline": "Icelandair",
                "return_airline": "Lufthansa",
                "stop_result_label": "1 Stop",
                "leg_durations": [500, 620],
            },
        ),
        SimpleNamespace(
            origin="YYZ",
            destination="BER",
            depart_date=date(2026, 5, 21),
            airline="Air Canada / LOT",
            price=851.0,
            stops=2,
            stop_label="2 Stop",
            itinerary_data={
                "return_date": "2026-06-01",
                "return_origin": "BUD",
                "outbound_airline": "Air Canada",
                "return_airline": "LOT",
                "stop_result_label": "2 Stop",
                "leg_durations": [540, 660],
            },
        ),
    ]

    workbook_bytes = export_route_group(group, results)
    workbook = load_workbook(BytesIO(workbook_bytes))

    assert "Toronto Open Jaw" in workbook.sheetnames
    assert workbook.sheetnames == ["Toronto Open Jaw"]

    sheet = workbook["Toronto Open Jaw"]
    headers = [sheet.cell(row=1, column=index).value for index in range(1, 11)]
    assert headers == [
        "Date",
        "Ending Date",
        "Dep Airport",
        "Arrival Airport",
        "Return From",
        "Nights",
        "Airline",
        "Stop Result",
        "Duration",
        "Flight Price",
    ]
    assert sheet["A2"].value == datetime(2026, 5, 20)
    assert sheet["B2"].value == datetime(2026, 5, 31)
    assert sheet["D2"].value == "BER"
    assert sheet["E2"].value == "BUD"  # Return From (return-leg origin)
    assert sheet["F2"].value == 11
    assert sheet["G2"].value == "Icelandair / Lufthansa"
    assert sheet["H2"].value == "1 Stop"
    assert sheet["I2"].value == "8h 20m / 10h 20m"
    assert sheet["J2"].value == 829


def test_multi_city_export_marks_missing_dates_as_na() -> None:
    group = SimpleNamespace(
        trip_type="multi_city",
        origins=["YYZ"],
        nights=11,
        days_ahead=3,
        start_date=date(2026, 5, 20),
        end_date=date(2026, 5, 22),
        sheet_name_map={"YYZ": "Toronto Open Jaw"},
    )

    results = [
        SimpleNamespace(
            origin="YYZ",
            destination="BER",
            depart_date=date(2026, 5, 20),
            airline="Icelandair / Lufthansa",
            price=829.0,
            stops=1,
            stop_label="1 Stop",
            itinerary_data={
                "return_date": "2026-05-31",
                "return_origin": "BUD",
                "outbound_airline": "Icelandair",
                "return_airline": "Lufthansa",
                "stop_result_label": "1 Stop",
                "leg_durations": [500, 620],
            },
        ),
    ]

    workbook_bytes = export_route_group(group, results)
    workbook = load_workbook(BytesIO(workbook_bytes))
    sheet = workbook["Toronto Open Jaw"]

    assert sheet["A2"].value == datetime(2026, 5, 20)
    assert sheet["A3"].value == datetime(2026, 5, 21)
    assert sheet["A4"].value == datetime(2026, 5, 22)
    assert sheet["B3"].value == "N-A"  # Ending Date
    assert sheet["E3"].value == "N-A"  # Return From
    assert sheet["G3"].value == "N-A"  # Airline
    assert sheet["H3"].value == "N-A"  # Stop Result
    assert sheet["I3"].value == "N-A"  # Duration
    assert sheet["J3"].value == "N-A"  # Flight Price
