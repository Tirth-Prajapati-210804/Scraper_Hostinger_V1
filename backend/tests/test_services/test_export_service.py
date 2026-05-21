from __future__ import annotations

import uuid
from datetime import date, timedelta
from io import BytesIO
from unittest.mock import MagicMock

import openpyxl
import pytest

from app.services.export_service import export_route_group


# ── helpers ──────────────────────────────────────────────────────────────────

def make_route_group(
    sheet_name_map: dict | None = None,
    special_sheets: list | None = None,
    destination_label: str = "SGN",
    nights: int = 7,
) -> MagicMock:
    rg = MagicMock()
    rg.id = uuid.uuid4()
    rg.name = "Test Group"
    rg.destination_label = destination_label
    rg.nights = nights
    rg.sheet_name_map = sheet_name_map or {"YVR": "YVR"}
    rg.special_sheets = special_sheets or []
    rg.trip_type = "one_way"
    return rg


def make_result(
    origin: str = "YVR",
    destination: str = "SGN",
    depart_date: date | None = None,
    price: float = 200.0,
    airline: str = "VJ",
) -> MagicMock:
    r = MagicMock()
    r.origin = origin
    r.destination = destination
    r.depart_date = depart_date or (date.today() + timedelta(days=1))
    r.price = price
    r.airline = airline
    r.stop_label = ""
    r.stops = 1
    r.duration_minutes = 120
    r.itinerary_data = None
    return r


# ── tests ─────────────────────────────────────────────────────────────────────

def test_export_creates_one_sheet_per_origin() -> None:
    rg = make_route_group(sheet_name_map={"YVR": "Vancouver", "YYZ": "Toronto"})
    results = [make_result(origin="YVR"), make_result(origin="YYZ")]
    wb = openpyxl.load_workbook(BytesIO(export_route_group(rg, results)))
    assert "Vancouver" in wb.sheetnames
    assert "Toronto" in wb.sheetnames


def test_export_has_correct_headers() -> None:
    rg = make_route_group()
    wb = openpyxl.load_workbook(BytesIO(export_route_group(rg, [make_result()])))
    ws = wb["YVR"]
    assert ws.cell(1, 1).value == "Date"
    assert ws.cell(1, 2).value == "Dep Airport"
    assert ws.cell(1, 3).value == "Arrival Airport"
    assert ws.cell(1, 4).value == "Nights"
    assert ws.cell(1, 5).value == "Airline"
    assert ws.cell(1, 6).value == "Stop Result"
    assert ws.cell(1, 7).value == "Duration"
    assert ws.cell(1, 8).value == "Flight Price"
    assert ws.cell(2, 1).number_format == "DD-MM-YYYY"


def test_export_destination_label_in_arrivel_column() -> None:
    rg = make_route_group(destination_label="TYO/SHA")
    result = make_result()
    wb = openpyxl.load_workbook(BytesIO(export_route_group(rg, [result])))
    ws = wb["YVR"]
    assert ws.cell(2, 3).value == "TYO/SHA"


def test_export_nights_in_night_column() -> None:
    rg = make_route_group(nights=12)
    result = make_result()
    wb = openpyxl.load_workbook(BytesIO(export_route_group(rg, [result])))
    ws = wb["YVR"]
    assert ws.cell(2, 4).value == 12


def test_export_prices_are_integers() -> None:
    rg = make_route_group()
    wb = openpyxl.load_workbook(BytesIO(export_route_group(rg, [make_result(price=199.75)])))
    ws = wb["YVR"]
    assert ws.cell(2, 7).value == "2h0min"
    assert ws.cell(2, 8).value == 200


def test_export_cheapest_per_date() -> None:
    rg = make_route_group()
    today = date.today()
    d = today + timedelta(days=1)
    results = [
        make_result(origin="YVR", destination="SGN", depart_date=d, price=500.0, airline="XX"),
        make_result(origin="YVR", destination="HAN", depart_date=d, price=300.0, airline="VN"),
    ]
    wb = openpyxl.load_workbook(BytesIO(export_route_group(rg, results)))
    ws = wb["YVR"]
    assert ws.cell(2, 8).value == 300
    assert ws.cell(2, 5).value == "VN"


def test_export_missing_date_shows_none_price() -> None:
    rg = make_route_group(sheet_name_map={"YVR": "YVR", "YYZ": "Toronto"})
    today = date.today()
    results = [make_result(origin="YVR", depart_date=today + timedelta(days=1), price=100.0)]
    wb = openpyxl.load_workbook(BytesIO(export_route_group(rg, results)))
    ws = wb["Toronto"]
    assert ws.cell(2, 8).value is None


def test_export_special_sheet_4_columns() -> None:
    special = {
        "name": "Special",
        "origin": "YVR",
        "destinations": ["SGN"],
        "destination_label": "SGN",
        "columns": 4,
    }
    rg = make_route_group(special_sheets=[special])
    result = make_result(origin="YVR", destination="SGN", price=250.0)
    wb = openpyxl.load_workbook(BytesIO(export_route_group(rg, [result])))
    ws = wb["Special"]
    assert ws.cell(1, 1).value == "Date"
    assert ws.cell(1, 2).value == "Dep Airport"
    assert ws.cell(1, 3).value == "Arrival Airport"
    assert ws.cell(1, 4).value == "Flight Price"
    assert ws.cell(2, 4).value == 250


def test_export_special_sheet_6_columns() -> None:
    special = {
        "name": "Multi",
        "origin": "YVR",
        "destinations": ["SGN", "HAN"],
        "destination_label": "VN",
        "columns": 6,
    }
    rg = make_route_group(nights=7, special_sheets=[special])
    result = make_result(origin="YVR", destination="SGN", price=400.0, airline="VN")
    wb = openpyxl.load_workbook(BytesIO(export_route_group(rg, [result])))
    ws = wb["Multi"]
    assert ws.cell(1, 4).value == "Nights"
    assert ws.cell(1, 5).value == "Airline"
    assert ws.cell(1, 6).value == "Stop Result"
    assert ws.cell(1, 7).value == "Duration"
    assert ws.cell(1, 8).value == "Flight Price"
    assert ws.cell(2, 4).value == 7
    assert ws.cell(2, 5).value == "VN"
    assert ws.cell(2, 8).value == 400


def test_export_uses_per_leg_duration_label_when_available() -> None:
    rg = make_route_group()
    result = make_result()
    result.duration_minutes = 2175
    result.itinerary_data = {"leg_durations": [1450, 725]}

    wb = openpyxl.load_workbook(BytesIO(export_route_group(rg, [result])))
    ws = wb["YVR"]

    assert ws.cell(2, 7).value == "24h10min / 12h5min"


def test_multi_city_export_creates_one_sheet_per_route() -> None:
    rg = make_route_group(sheet_name_map={"YOW": "YOW"})
    rg.trip_type = "multi_city"
    rg.origins = ["YOW"]

    first = make_result(origin="YOW", destination="LHR", price=671.0, airline="Air France")
    first.itinerary_data = {
        "return_date": "2026-06-13",
        "return_origin": "MXP",
    }

    second = make_result(origin="YOW", destination="LGW", price=702.0, airline="KLM")
    second.itinerary_data = {
        "return_date": "2026-06-13",
        "return_origin": "MXP",
    }

    wb = openpyxl.load_workbook(BytesIO(export_route_group(rg, [first, second])))

    assert wb.sheetnames == ["YOW-LGW", "YOW-LHR"]
    assert wb["YOW-LHR"].cell(1, 7).value == "Stop Result"
    assert wb["YOW-LHR"].cell(1, 8).value == "Duration"
    assert wb["YOW-LHR"].cell(1, 9).value == "Flight Price"
    assert wb["YOW-LHR"].cell(2, 4).value == "LHR"
    assert wb["YOW-LGW"].cell(2, 4).value == "LGW"
    assert wb["YOW-LHR"].cell(2, 1).number_format == "DD-MM-YYYY"
    assert wb["YOW-LHR"].cell(2, 2).number_format == "DD-MM-YYYY"
